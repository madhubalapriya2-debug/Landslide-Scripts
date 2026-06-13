
# =========================================================
# LITHOLOGY RECLASSIFICATION SCRIPT
# =========================================================

import os

from qgis.core import (
    QgsRasterLayer,
    QgsProject
)

from qgis.analysis import (
    QgsRasterCalculator,
    QgsRasterCalculatorEntry
)

from openpyxl import load_workbook


# =========================================================
# INPUTS
# =========================================================

base = "/Users/madhubalapriya/Downloads/Kerala/processed/Model_Input"

excel_path = os.path.join(base, "risk.xlsx")

lithology_raster = os.path.join(base, "Lithology.tif")

out_folder = os.path.join(base, "Reclassified")

os.makedirs(out_folder, exist_ok=True)

output_raster = os.path.join(
    out_folder,
    "Lithology_Reclass.tif"
)


# =========================================================
# LOAD LITHOLOGY RULES
# =========================================================
def read_lithology_rules(workbook_path, sheet_name):

    wb = load_workbook(workbook_path, data_only=True)

    print("Available Sheets:")
    print(wb.sheetnames)

    if sheet_name not in wb.sheetnames:

        raise Exception(
            f"Sheet '{sheet_name}' not found"
        )

    ws = wb[sheet_name]

    # -----------------------------------------------------
    # READ HEADER
    # -----------------------------------------------------
    header = [
        str(c.value).strip()
        if c.value is not None else ""
        for c in ws[1]
    ]

    print("\nLithology Headers:")
    print(header)

    header_lower = [
        h.lower().strip()
        for h in header
    ]

    # -----------------------------------------------------
    # COLUMN INDEX
    # -----------------------------------------------------
    value_idx = header_lower.index("value") + 1
    risk_idx = header_lower.index("riskscore") + 1

    # -----------------------------------------------------
    # READ RULES
    # -----------------------------------------------------
    rules = {}

    for row in range(2, ws.max_row + 1):

        lithology = ws.cell(row, value_idx).value
        risk = ws.cell(row, risk_idx).value

        if lithology is None or risk is None:
            continue

        lithology_name = str(lithology).strip().upper()

        rules[lithology_name] = int(risk)

    return rules


# =========================================================
# VALIDATE RULES
# =========================================================
def validate_rules(rules):

    if not rules:

        raise Exception(
            "No Lithology rules found"
        )

    for lithology, risk in rules.items():

        if risk not in [1, 2, 3, 4, 5]:

            raise Exception(
                f"Invalid RiskScore {risk} "
                f"for lithology '{lithology}'"
            )


# =========================================================
# LOAD RASTER
# =========================================================
def load_raster(path, name):

    if not os.path.exists(path):

        raise Exception(
            f"{name} raster not found:\n{path}"
        )

    layer = QgsRasterLayer(path, name)

    if not layer.isValid():

        raise Exception(
            f"{name} raster invalid"
        )

    return layer


# =========================================================
# BUILD RECLASS EXPRESSION
# =========================================================
def build_expression(ref_name, rules):

    """
    IMPORTANT:
    Raster values must match lithology class IDs.

    Example:

    1 = MICA SCHIST
    2 = GNEISS
    3 = GRANULITE
    etc.
    """

    parts = []

    lithology_code = 1

    for lithology_name, risk in rules.items():

        print(
            f"Class {lithology_code} -> "
            f"{lithology_name} -> Risk {risk}"
        )

        parts.append(
            f'(({ref_name}@1 = {lithology_code}) * {risk})'
        )

        lithology_code += 1

    expression = " + ".join(parts)

    return expression


# =========================================================
# RUN RECLASSIFICATION
# =========================================================
def run_reclassification(
    layer,
    layer_name,
    expression,
    output_path
):

    entry = QgsRasterCalculatorEntry()

    entry.ref = f"{layer_name}@1"
    entry.raster = layer
    entry.bandNumber = 1

    calc = QgsRasterCalculator(
        expression,
        output_path,
        "GTiff",
        layer.extent(),
        layer.width(),
        layer.height(),
        [entry]
    )

    result = calc.processCalculation()

    if result == 0:

        print(
            f"\n{layer_name} reclassified successfully"
        )

        out_layer = QgsRasterLayer(
            output_path,
            f"{layer_name}_Reclass"
        )

        if out_layer.isValid():

            QgsProject.instance().addMapLayer(
                out_layer
            )

            print(
                f"Output added to QGIS:\n{output_path}"
            )

    else:

        raise Exception(
            f"{layer_name} reclassification failed"
        )


# =========================================================
# MAIN
# =========================================================
print("\n====================================")
print("STARTING LITHOLOGY RECLASSIFICATION")
print("====================================")

# ---------------------------------------------------------
# LOAD RULES
# ---------------------------------------------------------
rules = read_lithology_rules(
    excel_path,
    "Lithology"
)

validate_rules(rules)

print("\nLithology Rules:")
print(rules)

# ---------------------------------------------------------
# LOAD RASTER
# ---------------------------------------------------------
layer = load_raster(
    lithology_raster,
    "Lithology"
)

QgsProject.instance().addMapLayer(layer)

# ---------------------------------------------------------
# BUILD EXPRESSION
# ---------------------------------------------------------
expression = build_expression(
    "Lithology",
    rules
)

print("\nRaster Calculator Expression:")
print(expression)

# ---------------------------------------------------------
# RUN
# ---------------------------------------------------------
run_reclassification(
    layer,
    "Lithology",
    expression,
    output_raster
)

print("\n====================================")
print("LITHOLOGY RECLASSIFICATION FINISHED")
print("====================================")

