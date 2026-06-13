import os
import numpy as np
from osgeo import gdal
from openpyxl import Workbook
from openpyxl.styles import Font

# --------------------------------------------------
# INPUT PATHS
# --------------------------------------------------
base = r"/Users/madhubalapriya/Downloads/Kerala/processed/Model_Input"

rasters = {
    "Elevation": os.path.join(base, "Elevation.tif"),
    "LULC": os.path.join(base, "LULC.tif"),
    "DistanceToRiver": os.path.join(base, "Distance_River.tif"),
    "Slope": os.path.join(base, "Slope.tif"),
    "Rainfall": os.path.join(base, "Rainfall.tif"),
    "NDVI" : os.path.join(base, "NDVI.tif"),
    "Aspect": os.path.join(base, "Aspect.tif"),
    "STI": os.path.join(base,"STI.tif"),
    "TWI": os.path.join(base,"TWI.tif")
}

# --------------------------------------------------
# OUTPUT EXCEL
# --------------------------------------------------
out_excel = os.path.join(base, "Reclass_Assessment.xlsx")

# --------------------------------------------------
# HELPERS
# --------------------------------------------------
def read_raster_array(raster_path):
    ds = gdal.Open(raster_path)

    if ds is None:
        raise Exception(f"Could not open raster: {raster_path}")

    band = ds.GetRasterBand(1)

    arr = band.ReadAsArray().astype("float64")

    nodata = band.GetNoDataValue()

    if nodata is not None:
        arr[arr == nodata] = np.nan

    # remove infinite values
    arr[~np.isfinite(arr)] = np.nan

    return arr, ds, nodata


def raster_summary(arr):
    valid = arr[~np.isnan(arr)]

    if valid.size == 0:
        return None

    return {
        "count_valid": int(valid.size),
        "min": float(np.min(valid)),
        "max": float(np.max(valid)),
        "mean": float(np.mean(valid)),
        "std": float(np.std(valid)),
        "q0": float(np.percentile(valid, 0)),
        "q10": float(np.percentile(valid, 10)),
        "q20": float(np.percentile(valid, 20)),
        "q25": float(np.percentile(valid, 25)),
        "q40": float(np.percentile(valid, 40)),
        "q50": float(np.percentile(valid, 50)),
        "q60": float(np.percentile(valid, 60)),
        "q75": float(np.percentile(valid, 75)),
        "q80": float(np.percentile(valid, 80)),
        "q90": float(np.percentile(valid, 90)),
        "q100": float(np.percentile(valid, 100)),
    }


def unique_value_counts(arr, max_unique_for_full=500):
    valid = arr[~np.isnan(arr)]

    if valid.size == 0:
        return []

    unique_vals, counts = np.unique(valid, return_counts=True)

    if len(unique_vals) > max_unique_for_full:
        return None

    results = []

    for val, cnt in zip(unique_vals, counts):
        results.append((float(val), int(cnt)))

    return results


def add_header(ws, row_num, headers):
    for col_num, h in enumerate(headers, start=1):
        cell = ws.cell(row=row_num, column=col_num, value=h)
        cell.font = Font(bold=True)


# --------------------------------------------------
# WORKBOOK SETUP
# --------------------------------------------------
wb = Workbook()

default_sheet = wb.active
wb.remove(default_sheet)

summary_ws = wb.create_sheet("Summary")

add_header(summary_ws, 1, [
    "Raster",
    "Path",
    "Valid_Count",
    "Min",
    "Max",
    "Mean",
    "StdDev",
    "Q10",
    "Q20",
    "Q25",
    "Q40",
    "Q50",
    "Q60",
    "Q75",
    "Q80",
    "Q90"
])

summary_row = 2

# --------------------------------------------------
# CONTINUOUS RASTERS
# --------------------------------------------------
continuous_for_quantiles = [
    "Elevation",
    "Rainfall",
    "TWI", 
    "NDVI",
    "STI",
    "Aspect", 
]

# --------------------------------------------------
# MAIN PROCESSING
# --------------------------------------------------
for name, path in rasters.items():

    if not os.path.exists(path):
        print(f"{name} missing: {path}")
        continue

    print(f"Processing: {name}")

    arr, ds, nodata = read_raster_array(path)

    stats = raster_summary(arr)

    if stats is None:
        print(f"No valid cells in {name}")
        continue

    # --------------------------------------------------
    # SUMMARY SHEET
    # --------------------------------------------------
    summary_ws.cell(summary_row, 1, name)
    summary_ws.cell(summary_row, 2, path)
    summary_ws.cell(summary_row, 3, stats["count_valid"])
    summary_ws.cell(summary_row, 4, stats["min"])
    summary_ws.cell(summary_row, 5, stats["max"])
    summary_ws.cell(summary_row, 6, stats["mean"])
    summary_ws.cell(summary_row, 7, stats["std"])
    summary_ws.cell(summary_row, 8, stats["q10"])
    summary_ws.cell(summary_row, 9, stats["q20"])
    summary_ws.cell(summary_row, 10, stats["q25"])
    summary_ws.cell(summary_row, 11, stats["q40"])
    summary_ws.cell(summary_row, 12, stats["q50"])
    summary_ws.cell(summary_row, 13, stats["q60"])
    summary_ws.cell(summary_row, 14, stats["q75"])
    summary_ws.cell(summary_row, 15, stats["q80"])
    summary_ws.cell(summary_row, 16, stats["q90"])

    summary_row += 1

    # --------------------------------------------------
    # QUANTILE SHEETS
    # --------------------------------------------------
    if name in continuous_for_quantiles:

        ws = wb.create_sheet(f"{name}_Quantiles")

        add_header(ws, 1, ["Metric", "Value"])

        quantile_rows = [
            ("Min", stats["min"]),
            ("Q10", stats["q10"]),
            ("Q20", stats["q20"]),
            ("Q25", stats["q25"]),
            ("Q40", stats["q40"]),
            ("Q50", stats["q50"]),
            ("Q60", stats["q60"]),
            ("Q75", stats["q75"]),
            ("Q80", stats["q80"]),
            ("Q90", stats["q90"]),
            ("Max", stats["max"]),
        ]

        r = 2

        for metric, value in quantile_rows:
            ws.cell(r, 1, metric)
            ws.cell(r, 2, value)
            r += 1

        # Suggested Classes
        r += 1

        add_header(ws, r, [
            "Suggested_Class",
            "From",
            "To",
            "Risk_Score"
        ])

        r += 1

        suggested = [
            ("Class 1", stats["min"], stats["q20"], ""),
            ("Class 2", stats["q20"], stats["q40"], ""),
            ("Class 3", stats["q40"], stats["q60"], ""),
            ("Class 4", stats["q60"], stats["q80"], ""),
            ("Class 5", stats["q80"], stats["max"], ""),
        ]

        for row in suggested:
            for c, val in enumerate(row, start=1):
                ws.cell(r, c, val)

            r += 1

        # logic note
        r += 1

        if name == "Elevation":
            ws.cell(r, 1, "Flood Logic")
            ws.cell(r, 2, "Lower elevation = higher flood risk")

        elif name == "Rainfall":
            ws.cell(r, 1, "Flood Logic")
            ws.cell(r, 2, "Higher rainfall = higher flood risk")

    # --------------------------------------------------
    # LULC CATEGORIES
    # --------------------------------------------------
    if name == "LULC":

        ws = wb.create_sheet("LULC_Categories")

        add_header(ws, 1, [
            "LULC_Value",
            "Pixel_Count",
            "Class_Name",
            "Flood_Risk_Score_1_to_5"
        ])

        unique_counts = unique_value_counts(
            arr,
            max_unique_for_full=1000
        )

        if unique_counts is None:

            ws.cell(
                2,
                1,
                "Too many unique values. Check whether raster is categorical."
            )

        else:
            r = 2

            for val, cnt in unique_counts:
                ws.cell(r, 1, val)
                ws.cell(r, 2, cnt)
                ws.cell(r, 3, "")
                ws.cell(r, 4, "")

                r += 1

    # --------------------------------------------------
    # SLOPE TEMPLATE
    # --------------------------------------------------
    if name == "Slope":

        ws = wb.create_sheet("Slope_Template")

        add_header(ws, 1, [
            "From_Degree",
            "To_Degree",
            "Risk_Score_1_to_5",
            "Reason"
        ])

        template = [
            (0, 2, "", "Very flat = more water stagnation"),
            (2, 5, "", "Low slope"),
            (5, 10, "", "Moderate slope"),
            (10, 20, "", "Steeper slope"),
            (20, 999999, "", "Very steep = less flood stagnation"),
        ]

        r = 2

        for row in template:
            for c, val in enumerate(row, start=1):
                ws.cell(r, c, val)

            r += 1

    # --------------------------------------------------
    # DISTANCE TO RIVER TEMPLATE
    # --------------------------------------------------
    if name == "DistanceToRiver":

        ws = wb.create_sheet("DistRiver_Template")

        add_header(ws, 1, [
            "From_Meter",
            "To_Meter",
            "Risk_Score_1_to_5",
            "Reason"
        ])

        template = [
            (0, 250, "", "Closest to river = highest flood influence"),
            (250, 500, "", "Near river"),
            (500, 1000, "", "Moderate influence"),
            (1000, 2000, "", "Low influence"),
            (2000, 999999, "", "Far from river = lowest influence"),
        ]

        r = 2

        for row in template:
            for c, val in enumerate(row, start=1):
                ws.cell(r, c, val)

            r += 1


# --------------------------------------------------
# AUTO COLUMN WIDTH
# --------------------------------------------------
for ws in wb.worksheets:

    for col in ws.columns:

        max_len = 0

        col_letter = col[0].column_letter

        for cell in col:

            try:
                val = str(cell.value) if cell.value is not None else ""

                max_len = max(max_len, len(val))

            except:
                pass

        ws.column_dimensions[col_letter].width = min(max_len + 2, 35)

# --------------------------------------------------
# SAVE EXCEL
# --------------------------------------------------
wb.save(out_excel)

print(f"\nExcel created successfully:\n{out_excel}")