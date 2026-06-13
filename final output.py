#PHASE2 

from qgis.core import (
    QgsProcessing,
    QgsProcessingAlgorithm,
    # QgsProcessingParameterFolderDestination,
    
    QgsProcessingParameterFile,
    QgsProcessingParameterNumber,
    QgsProcessingException,
    QgsRasterLayer,
    QgsProcessingParameterEnum,
    QgsProcessingParameterString
)
from qgis.analysis import QgsRasterCalculator, QgsRasterCalculatorEntry
from openpyxl import load_workbook
import os
import numpy as np
from osgeo import gdal

class FloodPhase2Tool(QgsProcessingAlgorithm):

    WORKSPACE = "WORKSPACE"
    RECLASS_EXCEL = "RECLASS_EXCEL"

    WT_RAINFALL = "WT_RAINFALL"
    WT_RIVER_DISTANCE = "WT_RIVER_DISTANCE"
    WT_ELEVATION = "WT_ELEVATION"
    WT_SLOPE = "WT_SLOPE"
    WT_ASPECT = "WT_ASPECT"
    WT_LULC = "WT_LULC"
    WT_NDVI = "WT_NDVI"
    WT_STI = "WT_STI"
    WT_TWI = "WT_TWI"
    WT_LITHOLOGY = "WT_LITHOLOGY"
    FINAL_CLASS_METHOD = "FINAL_CLASS_METHOD"
    FINAL_CLASS_COUNT = "FINAL_CLASS_COUNT"
    MANUAL_BREAKS = "MANUAL_BREAKS"
    def name(self):
        return "flood_phase3"

    def displayName(self):
        return "Landslide Estimation Custom Reclass"

    def group(self):
        return "Landslide Tools"

    def groupId(self):
        return "Landslide_tools"

    def shortHelpString(self):
        return (
            "Reads Phase 1 rasters, applies reclassification from Excel, "
            "and generates weighted landslide susceptibility index."
        )

    def initAlgorithm(self, config=None):
        self.addParameter(
        #     QgsProcessingParameterFolderDestination(
        #         self.WORKSPACE,
        #         "Phase 1 Workspace Folder"
        #     )
      
            QgsProcessingParameterFile(
                self.WORKSPACE,
                "Phase 1 Workspace Folder",
                behavior=QgsProcessingParameterFile.Folder
            )

        )
        self.addParameter(
            QgsProcessingParameterFile(
                self.RECLASS_EXCEL,
                "Reclassification Excel File",
                extension="xlsx"
            )
        )

        self.addParameter(
            QgsProcessingParameterNumber(
                self.WT_RAINFALL,
                "Weight - Rainfall",
                type=QgsProcessingParameterNumber.Double,
                defaultValue=0.203

            )
        )

        self.addParameter(
            QgsProcessingParameterNumber(
                self.WT_RIVER_DISTANCE,
                "Weight - River Distance",
                type=QgsProcessingParameterNumber.Double,
                defaultValue=0.076

            )
        )

        self.addParameter(
            QgsProcessingParameterNumber(
                self.WT_ELEVATION,
                "Weight - Elevation",
                type=QgsProcessingParameterNumber.Double,
                defaultValue=0.053

            )
        )

        self.addParameter(
            QgsProcessingParameterNumber(
                self.WT_SLOPE,
                "Weight - Slope",
                type=QgsProcessingParameterNumber.Double,
                defaultValue=0.218

            )
        )
        self.addParameter(
            QgsProcessingParameterNumber(
                self.WT_ASPECT,
                "Weight - Aspect",
                type=QgsProcessingParameterNumber.Double,
                defaultValue=0.064
           )
        )

        self.addParameter(
            QgsProcessingParameterNumber(
                self.WT_LULC,
                "Weight - LULC",
                type=QgsProcessingParameterNumber.Double,
                defaultValue=0.083

            )
        )

        self.addParameter(
            QgsProcessingParameterNumber(
                self.WT_NDVI,
                "Weight - NDVI",
                type=QgsProcessingParameterNumber.Double,
                defaultValue=0.069

            )
        )
        self.addParameter(
            QgsProcessingParameterNumber(
                self.WT_STI,
                "Weight - STI",
                type=QgsProcessingParameterNumber.Double,
                defaultValue=0.051

           )
         )
        self.addParameter(
            QgsProcessingParameterNumber(
                self.WT_TWI,
                "Weight - TWI",
                type=QgsProcessingParameterNumber.Double,
                defaultValue=0.138
            )
        )
        self.addParameter(
            QgsProcessingParameterNumber(
                self.WT_LITHOLOGY,
                "Weight - LITHOLOGY",
                type=QgsProcessingParameterNumber.Double,
                defaultValue=0.518
           )
        )

        self.addParameter(
            QgsProcessingParameterEnum(
                self.FINAL_CLASS_METHOD,
                "Final FEI Classification Method",
                options=["None", "Quantile", "Manual Breaks"],
                defaultValue=1
            )
        )

        self.addParameter(
            QgsProcessingParameterNumber(
                self.FINAL_CLASS_COUNT,
                "Number of Final Classes",
                type=QgsProcessingParameterNumber.Integer,
                defaultValue=5,
                optional=True
            )
        )

        self.addParameter(
            QgsProcessingParameterString(
                self.MANUAL_BREAKS,
                "Manual Break Values (comma separated, e.g. 0.2,0.35,0.5,0.7)",
                optional=True
            )
        )

    # ---------------------------------------------------
    # HELPERS
    # ---------------------------------------------------
    def ensure_folder(self, path):
        os.makedirs(path, exist_ok=True)
        return path

    def load_raster(self, path, name):
        if not os.path.exists(path):
            raise QgsProcessingException(f"{name} raster file not found: {path}")

        layer = QgsRasterLayer(path, name)
        if not layer.isValid():
            raise QgsProcessingException(f"{name} raster is invalid: {path}")

        return layer

    def read_excel_continuous_rules(self, workbook_path, sheet_name):
        wb = load_workbook(workbook_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            raise QgsProcessingException(f"Sheet '{sheet_name}' not found in Excel.")

        ws = wb[sheet_name]
        header = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]

        required = ["From", "To", "RiskScore"]
        for col in required:
            if col not in header:
                raise QgsProcessingException(f"Sheet '{sheet_name}' must contain columns: {required}")

        from_idx = header.index("From") + 1
        to_idx = header.index("To") + 1
        risk_idx = header.index("RiskScore") + 1

        rules = []
        for row in range(2, ws.max_row + 1):
            from_val = ws.cell(row, from_idx).value
            to_val = ws.cell(row, to_idx).value
            risk_val = ws.cell(row, risk_idx).value

            if from_val is None or to_val is None or risk_val is None:
                continue

            rules.append((float(from_val), float(to_val), int(risk_val)))

        return rules

    def read_excel_lulc_rules(self, workbook_path, sheet_name):
        wb = load_workbook(workbook_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            raise QgsProcessingException(f"Sheet '{sheet_name}' not found in Excel.")

        ws = wb[sheet_name]
        header = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]

        required = ["Value", "RiskScore"]
        for col in required:
            if col not in header:
                raise QgsProcessingException(f"Sheet '{sheet_name}' must contain columns: {required}")

        value_idx = header.index("Value") + 1
        risk_idx = header.index("RiskScore") + 1

        rules = {}
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row, value_idx).value
            risk_val = ws.cell(row, risk_idx).value

            if val is None or risk_val is None:
                continue

            rules[int(val)] = int(risk_val)

        return rules

    def validate_continuous_rules(self, rules, raster_name):
        if not rules:
            raise QgsProcessingException(f"No rules found for {raster_name}")

        for i, (from_val, to_val, risk) in enumerate(rules):
            if from_val >= to_val:
                raise QgsProcessingException(f"{raster_name}: invalid range {from_val} to {to_val}")
            if risk not in [1, 2, 3, 4, 5]:
                raise QgsProcessingException(f"{raster_name}: invalid RiskScore {risk}")

            if i > 0:
                prev_to = rules[i - 1][1]
                if from_val != prev_to:
                    # Warn only
                    pass

    def validate_lulc_rules(self, rules):
        if not rules:
            raise QgsProcessingException("No LULC rules found")

        for val, risk in rules.items():
            if risk not in [1, 2, 3, 4, 5]:
                raise QgsProcessingException(f"LULC: invalid RiskScore {risk} for value {val}")

    def build_continuous_expression(self, ref_name, rules):
        parts = []
        for i, (from_val, to_val, risk) in enumerate(rules):
            if i == len(rules) - 1:
                parts.append(f'(({ref_name}@1 >= {from_val}) AND ({ref_name}@1 <= {to_val})) * {risk}')
            else:
                parts.append(f'(({ref_name}@1 >= {from_val}) AND ({ref_name}@1 < {to_val})) * {risk}')
        return " + ".join(parts)

    def build_lulc_expression(self, ref_name, rules):
        parts = []
        for val, risk in rules.items():
            parts.append(f'(({ref_name}@1 = {val}) * {risk})')
        return " + ".join(parts)

    def run_reclass(self, layer, layer_name, expression, output_path):
        entry = QgsRasterCalculatorEntry()
        entry.ref = f"{layer_name}@1"
        entry.raster = layer
        entry.bandNumber = 1

        calc = QgsRasterCalculator(
            expression,
            output_path,
            "GTiff",
            layer.extent(),
            layer.width(),
            layer.height(),
            [entry]
        )

        result = calc.processCalculation()
        if result != 0:
            raise QgsProcessingException(f"{layer_name} reclassification failed with code: {result}")

        out_layer = QgsRasterLayer(output_path, f"{layer_name}_Reclass")
        if not out_layer.isValid():
            raise QgsProcessingException(f"{layer_name}_Reclass was created but could not be loaded.")

        return out_layer

    def run_weighted_overlay(self, loaded_layers, weights, output_path):
        entries = []

        for name, layer in loaded_layers.items():
            entry = QgsRasterCalculatorEntry()
            entry.ref = f"{name}@1"
            entry.raster = layer
            entry.bandNumber = 1
            entries.append(entry)

        ref_layer = loaded_layers["Rainfall"]

        expression = f"""
        (
       ("Lithology@1" * {weights['Lithology']}) +
       ("Distancefromriver@1" * {weights['Distancefromriver']}) +
       ("LULC@1" * {weights['LULC']}) +
       ("NDVI@1" * {weights['NDVI']}) +
       ("STI@1" * {weights['STI']}) +
       ("TWI@1" * {weights['TWI']}) +
       ("DEM@1" * {weights['DEM']}) +
       ("Aspect@1" * {weights['Aspect']}) +
       ("Slope@1" * {weights['Slope']}) +
       ("Rainfall@1" * {weights['Rainfall']})
        )
        """

        calc = QgsRasterCalculator(
            expression,
            output_path,
            "GTiff",
            ref_layer.extent(),
            ref_layer.width(),
            ref_layer.height(),
            entries
        )

        result = calc.processCalculation()
        if result != 0:
            raise QgsProcessingException(f"Weighted overlay failed with code: {result}")

        out_layer = QgsRasterLayer(output_path, "Landslide_Susceptibility_Index")
        if not out_layer.isValid():
            raise QgsProcessingException("Landslide susceptibility raster was created but could not be loaded.")

        return out_layer


    def classify_fei_quantile(self, fei_path, classified_out, num_classes, feedback):
        """
        Reclassify FEI raster into quantile-based classes.
        """
        ds = gdal.Open(fei_path)
        if ds is None:
            raise QgsProcessingException(f"Could not open raster: {fei_path}")

        band = ds.GetRasterBand(1)
        arr = band.ReadAsArray().astype("float64")

        nodata = band.GetNoDataValue()
        if nodata is not None:
            arr[arr == nodata] = np.nan

        arr[~np.isfinite(arr)] = np.nan
        valid = arr[~np.isnan(arr)]

        if valid.size == 0:
            raise QgsProcessingException("No valid raster values found in FEI raster.")

        # Example for 5 classes => [20, 40, 60, 80]
        percent_steps = [100 * i / num_classes for i in range(1, num_classes)]
        breaks = [float(np.percentile(valid, p)) for p in percent_steps]

        feedback.pushInfo(f"Quantile breaks: {breaks}")

        classified = np.full(arr.shape, -9999, dtype=np.int16)
        mask = ~np.isnan(arr)

        prev = None
        for i, brk in enumerate(breaks, start=1):
            if prev is None:
                classified[(mask) & (arr <= brk)] = i
            else:
                classified[(mask) & (arr > prev) & (arr <= brk)] = i
            prev = brk

        classified[(mask) & (arr > breaks[-1])] = num_classes

        self.save_classified_raster(ds, classified, classified_out)
        feedback.pushInfo(f"Quantile classified raster saved: {classified_out}")
        return classified_out


    def classify_fei_manual(self, fei_path, classified_out, manual_breaks, feedback):
        """
        Reclassify FEI raster using user-provided manual break values.
        Example manual_breaks: [0.2, 0.35, 0.5, 0.7]
        Produces 5 classes.
        """
        ds = gdal.Open(fei_path)
        if ds is None:
            raise QgsProcessingException(f"Could not open raster: {fei_path}")

        band = ds.GetRasterBand(1)
        arr = band.ReadAsArray().astype("float64")

        nodata = band.GetNoDataValue()
        if nodata is not None:
            arr[arr == nodata] = np.nan

        arr[~np.isfinite(arr)] = np.nan
        valid = arr[~np.isnan(arr)]

        if valid.size == 0:
            raise QgsProcessingException("No valid raster values found in FEI raster.")

        breaks = sorted(manual_breaks)
        feedback.pushInfo(f"Manual breaks: {breaks}")

        classified = np.full(arr.shape, -9999, dtype=np.int16)
        mask = ~np.isnan(arr)

        prev = None
        for i, brk in enumerate(breaks, start=1):
            if prev is None:
                classified[(mask) & (arr <= brk)] = i
            else:
                classified[(mask) & (arr > prev) & (arr <= brk)] = i
            prev = brk

        classified[(mask) & (arr > breaks[-1])] = len(breaks) + 1

        save_classified_raster(ds, classified, classified_out)
        feedback.pushInfo(f"Manual classified raster saved: {classified_out}")
        return classified_out


    def save_classified_raster(self, src_ds, classified_array, output_path):
        """
        Save classified raster preserving georeference.
        """
        driver = gdal.GetDriverByName("GTiff")
        out_ds = driver.Create(
            output_path,
            src_ds.RasterXSize,
            src_ds.RasterYSize,
            1,
            gdal.GDT_Int16
        )

        out_ds.SetGeoTransform(src_ds.GetGeoTransform())
        out_ds.SetProjection(src_ds.GetProjection())

        out_band = out_ds.GetRasterBand(1)
        out_band.WriteArray(classified_array)
        out_band.SetNoDataValue(-9999)
        out_band.FlushCache()

        out_ds = None
    # ---------------------------------------------------
    # MAIN
    # ---------------------------------------------------
    def processAlgorithm(self, parameters, context, feedback):
        workspace = self.parameterAsString(parameters, self.WORKSPACE, context)
        excel_path = self.parameterAsFile(parameters, self.RECLASS_EXCEL, context)

        weights = {
               "LULC": self.parameterAsDouble(parameters, self.WT_LULC, context),
               "NDVI": self.parameterAsDouble(parameters, self.WT_NDVI, context),
               "STI": self.parameterAsDouble(parameters, self.WT_STI, context),
                "TWI": self.parameterAsDouble(parameters, self.WT_TWI, context),
               "DEM": self.parameterAsDouble(parameters, self.WT_DEM, context),
                "Aspect": self.parameterAsDouble(parameters, self.WT_ASPECT, context),
               "Slope": self.parameterAsDouble(parameters, self.WT_SLOPE, context),
               "Rainfall": self.parameterAsDouble(parameters, self.WT_RAINFALL, context),
               "Lithology": self.parameterAsDouble(parameters, self.WT_LITHOLOGY, context),
               "Elevation": self.parameterAsDouble(parameters, self.WT_ELEVATION, context)

        }

        weight_sum = sum(weights.values())
        if abs(weight_sum - 1.0) > 0.001:
            raise QgsProcessingException(f"Weights must sum to 1. Current sum: {weight_sum}")

        model_input_folder = os.path.join(workspace, "04_model_input")
        parameter_folder = os.path.join(workspace, "05_parameters")
        reclass_folder = self.ensure_folder(os.path.join(workspace, "07_reclassified"))
        final_folder = self.ensure_folder(os.path.join(workspace, "08_final"))

        rasters = {
             "LULC": os.path.join(model_input_folder, "LULC.tif"),
             "NDVI": os.path.join(model_input_folder, "NDVI.tif"),
             "STI": os.path.join(model_input_folder, "STI.tif"),
             "TWI": os.path.join(model_input_folder, "TWI.tif"),
             "DEM": os.path.join(model_input_folder, "DEM.tif"),
             "Aspect": os.path.join(parameter_folder, "Aspect.tif"),
             "Slope": os.path.join(parameter_folder, "Slope.tif"),
             "Rainfall": os.path.join(model_input_folder, "Rainfall.tif"),
              "Lithology": os.path.join(model_input_folder, "Lithology.tif"),
              "Elevation": os.path.join(model_input_folder, "Elevation.tif")

        }

        continuous_names = [
            "Elevation",
            "Slope",
            "Aspect",
            "Distance_River",
            "Rainfall",
            "LULC",
            "NDVI",
            "TWI",
            "STI",
            "Lithology"
        ]

        feedback.pushInfo("Starting Phase 2: Reclassification + Weighted Overlay")

        reclassified_layers = {}

        for name, path in rasters.items():
            feedback.pushInfo(f"Processing: {name}")
            layer = self.load_raster(path, name)
            out_path = os.path.join(reclass_folder, f"{name}_Reclass.tif")

            if name in continuous_names:
                rules = self.read_excel_continuous_rules(excel_path, name)
                self.validate_continuous_rules(rules, name)
                expr = self.build_continuous_expression(name, rules)
                out_layer = self.run_reclass(layer, name, expr, out_path)
                reclassified_layers[name] = out_layer

            elif name == "LULC":
                rules = self.read_excel_lulc_rules(excel_path, "LULC")
                self.validate_lulc_rules(rules)
                expr = self.build_lulc_expression(name, rules)
                out_layer = self.run_reclass(layer, name, expr, out_path)
                reclassified_layers[name] = out_layer

        weighted_inputs = {
                "LULC": reclassified_layers["LULC"],
               "NDVI": reclassified_layers["NDVI"],
                "STI": reclassified_layers["STI"],
               "TWI": reclassified_layers["TWI"],
               "DEM": reclassified_layers["DEM"],
               "Aspect": reclassified_layers["Aspect"],
               "Slope": reclassified_layers["Slope"],
              "Rainfall": reclassified_layers["Rainfall"],
              "Lithology": reclassified_layers["Lithology"],
              "Elevation": reclassified_layers["Elevation"]

        }

        final_output = os.path.join(final_folder, "Landslide_Susceptibility_Index.tif")
        out_layer = self.run_weighted_overlay(weighted_inputs, weights, final_output)

        final_class_method = self.parameterAsEnum(parameters, self.FINAL_CLASS_METHOD, context)
        final_class_count = self.parameterAsInt(parameters, self.FINAL_CLASS_COUNT, context)
        manual_breaks_text = self.parameterAsString(parameters, self.MANUAL_BREAKS, context)

        classified_output = None

        if final_class_method == 1:  # Quantile
            classified_output = os.path.join(final_folder, f"Landslide_Susceptibility_{final_class_count}Class_Quantile.tif")
            self.classify_fei_quantile(
                final_output,
                classified_output,
                final_class_count,
                feedback
            )

        elif final_class_method == 2:  # Manual Breaks
            if not manual_breaks_text.strip():
                raise QgsProcessingException("Manual Breaks selected, but no break values were provided.")

            try:
                manual_breaks = [float(x.strip()) for x in manual_breaks_text.split(",") if x.strip()]
            except Exception:
                raise QgsProcessingException("Manual break values could not be parsed. Example: 0.2,0.35,0.5,0.7")

            classified_output = os.path.join(final_folder, "Lndslide_Susceptibility_ManualClass.tif")
            self.classify_fei_manual(
                final_output,
                classified_output,
                manual_breaks,
                feedback
            )
        feedback.pushInfo("Phase 2 completed successfully.")

        return {
            "RECLASS_FOLDER": reclass_folder,
            "FINAL_OUTPUT": final_output,
             "FINAL_CLASSIFIED_OUTPUT": classified_output if classified_output else ""
        }

    def createInstance(self):
        return LandslidePhase2Tool()
