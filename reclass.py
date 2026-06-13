# =========================================================
# RECLASSIFICATION SCRIPT
# =========================================================

import os
from qgis.core import (
    QgsRasterLayer,
    QgsVectorLayer,
    QgsProject
)

from qgis.analysis import (
    QgsRasterCalculator,
    QgsRasterCalculatorEntry
)

import processing

# =========================================================
# INPUTS
# =========================================================
base = "/Users/madhubalapriya/Downloads/Kerala/processed/Model_Input"

out_folder = os.path.join(base, "Reclassified")
os.makedirs(out_folder, exist_ok=True)

# ---------------------------------------------------------
# RULE SOURCE
# ---------------------------------------------------------
use_excel = True

excel_path = os.path.join(base, "risk.xlsx")

# csv_folder = os.path.join(base, "Reclass_CSV")

# =========================================================
# INPUT DATA
# =========================================================
rasters = {
    "elevation": os.path.join(base, "elevation.tif"),
    "Aspect": os.path.join(base, "Aspect.tif"),
    "slope": os.path.join(base, "slope.tif"),
    "Distance_River": os.path.join(base, "Distance_River.tif"),
    "Rainfall": os.path.join(base, "Rainfall.tif"),
    "LULC": os.path.join(base, "lulc.tif"),
    "NDVI": os.path.join(base, "NDVI.tif"),
    "STI": os.path.join(base, "STI.tif"),
    "TWI": os.path.join(base, "TWI.tif"),
    "Lithology": os.path.join(base, "Lithology.tif")
}


# =========================================================
# CONTINUOUS RASTERS
# =========================================================
continuous_names = [
    "elevation",
    "Distance_River",
    "slope",
    "TWI",
    "STI",
    "Aspect",
    "NDVI",
    "Lithology",
    "Rainfall"
]

# =========================================================
# READ CONTINUOUS RULES
# =========================================================
def read_excel_continuous_rules(workbook_path, sheet_name):

    from openpyxl import load_workbook

    wb = load_workbook(workbook_path, data_only=True)

    # Print available sheets
    print("Available Excel Sheets:")
    print(wb.sheetnames)

    # Check sheet existence
    if sheet_name not in wb.sheetnames:
        raise Exception(
            f"Sheet '{sheet_name}' not found in Excel.\n"
            f"Available sheets: {wb.sheetnames}"
        )

    ws = wb[sheet_name]

    rules = []

    header = [
        str(c.value).strip()
        if c.value is not None else ""
        for c in ws[1]
    ]

    from_idx = header.index("From") + 1
    to_idx = header.index("To") + 1
    risk_idx = header.index("RiskScore") + 1

    for row in range(2, ws.max_row + 1):

        from_val = ws.cell(row, from_idx).value
        to_val = ws.cell(row, to_idx).value
        risk_val = ws.cell(row, risk_idx).value

        # Skip empty rows
        if (
            from_val is None or str(from_val).strip() == "" or
            to_val is None or str(to_val).strip() == "" or
            risk_val is None or str(risk_val).strip() == ""
        ):
            continue

        rules.append((
            float(from_val),
            float(to_val),
            int(risk_val)
        ))

    return rules


# =========================================================
# READ LULC RULES
# =========================================================
def read_excel_lulc_rules(workbook_path, sheet_name):

    from openpyxl import load_workbook

    wb = load_workbook(workbook_path, data_only=True)
    ws = wb[sheet_name]

    rules = {}

    header = [
        str(c.value).strip()
        if c.value is not None else ""
        for c in ws[1]
    ]

    value_idx = header.index("Value") + 1
    risk_idx = header.index("RiskScore") + 1

    for row in range(2, ws.max_row + 1):

        val = ws.cell(row, value_idx).value
        risk_val = ws.cell(row, risk_idx).value

        if val is None or risk_val is None:
            continue

        rules[int(val)] = int(risk_val)

    return rules


# =========================================================
# READ LITHOLOGY RULES
# =========================================================
# =========================================================
# READ LITHOLOGY RULES
# =========================================================
def read_excel_Lithology_rules(workbook_path, sheet_name):

    from openpyxl import load_workbook

    wb = load_workbook(workbook_path, data_only=True)
    ws = wb[sheet_name]

    rules = {}

    header = [
        str(c.value).strip()
        if c.value is not None else ""
        for c in ws[1]
    ]

    lithology_idx = header.index("Value") + 1
    risk_idx = header.index("Riskscore") + 1

    for row in range(2, ws.max_row + 1):

        lithology = ws.cell(row, lithology_idx).value
        risk_val = ws.cell(row, risk_idx).value

        if lithology is None or risk_val is None:
            continue

        Lithology_name = str(Lithology).strip().upper()

        rules[Lithology_name] = int(risk_val)

    return rules
        
# =========================================================
# VALIDATION
# =========================================================
def validate_continuous_rules(rules, raster_name):

    if not rules:
        raise Exception(f"No rules found for {raster_name}")

    for i, (from_val, to_val, risk) in enumerate(rules):

        if from_val >= to_val:
            raise Exception(
                f"{raster_name}: invalid range {from_val} to {to_val}"
            )

        if risk not in [1, 2, 3, 4, 5]:
            raise Exception(
                f"{raster_name}: invalid RiskScore {risk}"
            )

        if i > 0:

            prev_to = rules[i - 1][1]

            if from_val != prev_to:
                print(
                    f"Warning: {raster_name} range gap/overlap "
                    f"near {prev_to} -> {from_val}"
                )


def validate_lulc_rules(rules):

    if not rules:
        raise Exception("No lulc rules found")

    for val, risk in rules.items():

        if risk not in [1, 2, 3, 4, 5]:
            raise Exception(
                f"lulc: invalid RiskScore {risk} for value {val}"
            )


def validate_Lithology_rules(rules):

    if not rules:
        raise Exception("No Lithology rules found")

    for Lithology, risk in rules.items():

        if risk not in [1, 2, 3, 4, 5]:
            raise Exception(
                f"Lithology: invalid RiskScore {risk} "
                f"for lithology '{lithology}'"
            )


# =========================================================
# LOAD RASTER
# =========================================================
def load_raster(path, name):

    if not os.path.exists(path):
        raise Exception(f"{name} raster not found: {path}")

    layer = QgsRasterLayer(path, name)

    if not layer.isValid():
        raise Exception(f"{name} raster invalid: {path}")

    return layer


# =========================================================
# BUILD EXPRESSIONS
# =========================================================
def build_continuous_expression(ref_name, rules):

    parts = []

    for from_val, to_val, risk in rules:

        parts.append(
            f'(({ref_name}@1 >= {from_val}) AND '
            f'({ref_name}@1 < {to_val})) * {risk}'
        )

    return " + ".join(parts)


def build_lulc_expression(ref_name, rules):

    parts = []

    for val, risk in rules.items():

        parts.append(
            f'(({ref_name}@1 = {val}) * {risk})'
        )

    return " + ".join(parts)


# =========================================================
# RUN RECLASSIFICATION
# =========================================================
def run_reclass(layer, layer_name, expression, output_path):

    entry = QgsRasterCalculatorEntry()
    entry.ref = f"{layer_name}@1"
    entry.raster = layer
    entry.bandNumber = 1

    calc = QgsRasterCalculator(
        expression,
        output_path,
        "GTiff",
        layer.extent(),
        layer.width(),
        layer.height(),
        [entry]
    )

    result = calc.processCalculation()

    if result == 0:

        print(f"{layer_name} reclassified successfully")

        out_layer = QgsRasterLayer(
            output_path,
            f"{layer_name}_Reclass"
        )

        if out_layer.isValid():
            QgsProject.instance().addMapLayer(out_layer)

    else:
        print(f"{layer_name} reclassification failed")


# =========================================================
# 

# =========================================================
# GET RULES
# =========================================================
def get_continuous_rules(name):

    if use_excel:
        return read_excel_continuous_rules(excel_path, name)

    else:
        raise Exception("CSV mode not implemented")


def get_lulc_rules():

    if use_excel:
        return read_excel_lulc_rules(excel_path, "lulc")

    else:
        raise Exception("CSV mode not implemented")


def get_Lithology_rules():

    if use_excel:
        return read_excel_Lithology_rules(
            excel_path,
            "Lithology"
        )

    else:
        raise Exception("CSV mode not implemented")


# =========================================================
# MAIN
# =========================================================
print("\n====================================")
print("STARTING RECLASSIFICATION")
print("====================================")

print(f"Reading rules from Excel: {excel_path}")

# =========================================================
# PROCESS RASTERS
# =========================================================
for name, path in rasters.items():

    print(f"\nProcessing: {name}")

    layer = load_raster(path, name)

    QgsProject.instance().addMapLayer(layer)

    out_path = os.path.join(
        out_folder,
        f"{name}_Reclass.tif"
    )

    # ---------------------------------
    # CONTINUOUS
    # ---------------------------------
    if name in continuous_names:

        rules = get_continuous_rules(name)

        validate_continuous_rules(rules, name)

        print("Rules:")
        print(rules)

        expr = build_continuous_expression(
            name,
            rules
        )

        print("Expression:")
        print(expr)

        run_reclass(
            layer,
            name,
            expr,
            out_path
        )

    # ---------------------------------
    # LULC
    # ---------------------------------
    elif name == "LULC":

        rules = get_lulc_rules()

        validate_lulc_rules(rules)

        print("LULC Rules:")
        print(rules)

        expr = build_lulc_expression(
            name,
            rules
        )

        print("Expression:")
        print(expr)

        run_reclass(
            layer,
            name,
            expr,
            out_path
        )

# =========================================================
# PROCESS LITHOLOGY
# =========================================================
print("\nProcessing: Lithology")

lithology_rules = get_Lithology_rules()

validate_lithology_rules(Lithology_rules)

print("Lithology Rules:")
print(Lithology_rules)

reclassify_Lithology_vector(
    Lithology_path,
    Lithology_rules
)

# =========================================================
# FINISHED
# =========================================================
print("\n====================================")
print("ALL RECLASSIFICATION COMPLETED")
print("====================================")